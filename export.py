# -*- coding: utf-8 -*-
"""
GetBeel 导出模块 / Export Module
提供多种格式导出功能 / Provides multiple format export functionality
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import config


class Exporter:
    """导出器类 / Exporter Class"""

    def __init__(self):
        """初始化导出器 / Initialize exporter"""
        self.storage = None

    def _get_storage(self):
        """获取存储实例 / Get storage instance"""
        if not self.storage:
            from storage import Storage
            self.storage = Storage()
        return self.storage

    def export_to_excel(self, output_path: Path, products: Optional[List[Dict[str, Any]]] = None) -> bool:
        """
        导出为 Excel 格式 / Export to Excel format

        Args:
            output_path: 输出文件路径 / Output file path
            products: 产品列表 / Products list

        Returns:
            是否导出成功 / Whether exported successfully
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("请安装 openpyxl 库: pip install openpyxl")
            return False

        if products is None:
            storage = self._get_storage()
            products = storage.get_products("today")

        if not products:
            print("没有产品数据可导出")
            return False

        # 创建工作簿 / Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Hunt Products"

        # 定义样式 / Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头 / Write headers
        headers = ["产品名称", "标语", "描述", "投票数", "评论数", "链接", "缩略图", "创作者", "发布日期", "话题"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据 / Write data
        for row, product in enumerate(products, 2):
            # 提取话题 / Extract topics
            topics_list = []
            topics = product.get('topics', {}).get('edges', [])
            for edge in topics:
                topic = edge.get('node', {})
                if topic.get('name'):
                    topics_list.append(topic['name'])

            # 提取创作者 / Extract makers
            maker = product.get('maker', {})
            maker_name = maker.get('name', 'Unknown')

            # 格式化日期 / Format date
            published_at = product.get('publishedAt', '')
            if published_at:
                try:
                    date = datetime.fromisoformat(published_at.replace('Z', '+00:00'))
                    published_at = date.strftime('%Y-%m-%d')
                except:
                    pass

            ws.cell(row=row, column=1, value=product.get('name', ''))
            ws.cell(row=row, column=2, value=product.get('tagline', ''))
            ws.cell(row=row, column=3, value=product.get('description', ''))
            ws.cell(row=row, column=4, value=product.get('votesCount', 0))
            ws.cell(row=row, column=5, value=product.get('commentsCount', 0))
            ws.cell(row=row, column=6, value=product.get('url', ''))
            ws.cell(row=row, column=7, value=product.get('thumbnail', {}).get('url', ''))
            ws.cell(row=row, column=8, value=maker_name)
            ws.cell(row=row, column=9, value=published_at)
            ws.cell(row=row, column=10, value=', '.join(topics_list))

        # 调整列宽 / Adjust column widths
        column_widths = [30, 40, 60, 10, 10, 50, 50, 20, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 保存文件 / Save file
        try:
            wb.save(output_path)
            print(f"Excel 导出成功: {output_path}")
            return True
        except Exception as e:
            print(f"Excel 保存失败: {e}")
            return False

    def export_to_markdown(self, output_path: Path, products: Optional[List[Dict[str, Any]]] = None) -> bool:
        """
        导出为 Markdown 格式 / Export to Markdown format

        Args:
            output_path: 输出文件路径 / Output file path
            products: 产品列表 / Products list

        Returns:
            是否导出成功 / Whether exported successfully
        """
        if products is None:
            storage = self._get_storage()
            products = storage.get_products("today")

        if not products:
            print("没有产品数据可导出")
            return False

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("# Product Hunt 热门产品\n\n")
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                f.write("---\n\n")

                for i, product in enumerate(products, 1):
                    f.write(f"## {i}. {product.get('name')}\n\n")
                    f.write(f"**标语**: {product.get('tagline')}\n\n")

                    if product.get('description'):
                        f.write(f"**描述**: {product.get('description')}\n\n")

                    maker = product.get('maker', {})
                    if maker:
                        f.write(f"**创作者**: {maker.get('name', 'Unknown')}\n\n")

                    f.write(f"- 投票: {product.get('votesCount', 0)}\n")
                    f.write(f"- 评论: {product.get('commentsCount', 0)}\n")
                    f.write(f"- [查看产品]({product.get('url')})\n\n")

                    # 话题 / Topics
                    topics = product.get('topics', {}).get('edges', [])
                    if topics:
                        topic_names = [t.get('node', {}).get('name', '') for t in topics if t.get('node')]
                        f.write(f"**话题**: {', '.join(topic_names)}\n\n")

                    f.write("---\n\n")

            print(f"Markdown 导出成功: {output_path}")
            return True
        except Exception as e:
            print(f"Markdown 导出失败: {e}")
            return False

    def export_to_html(self, output_path: Path, products: Optional[List[Dict[str, Any]]] = None) -> bool:
        """
        导出为 HTML 格式 / Export to HTML format

        Args:
            output_path: 输出文件路径 / Output file path
            products: 产品列表 / Products list

        Returns:
            是否导出成功 / Whether exported successfully
        """
        if products is None:
            storage = self._get_storage()
            products = storage.get_products("today")

        if not products:
            print("没有产品数据可导出")
            return False

        try:
            html = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Product Hunt 热门产品</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 900px; margin: 0 auto; padding: 20px; }
        h1 { color: #333; border-bottom: 2px solid #4472C4; padding-bottom: 10px; }
        .product { border: 1px solid #ddd; border-radius: 8px; padding: 20px; margin-bottom: 20px; }
        .product-name { font-size: 24px; color: #4472C4; margin: 0 0 10px 0; }
        .product-tagline { font-size: 16px; color: #666; margin: 10px 0; }
        .product-description { color: #555; margin: 10px 0; }
        .product-stats { display: flex; gap: 20px; margin: 15px 0; }
        .stat { background: #f5f5f5; padding: 8px 16px; border-radius: 4px; }
        .product-meta { color: #888; font-size: 14px; margin: 10px 0; }
        .product-link a { display: inline-block; background: #4472C4; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; }
        .product-link a:hover { background: #3567b5; }
        .topics { margin-top: 10px; }
        .topic { display: inline-block; background: #e8f4fd; color: #4472C4; padding: 4px 12px; border-radius: 12px; font-size: 12px; margin-right: 8px; }
    </style>
</head>
<body>
    <h1>Product Hunt 热门产品</h1>
    <p>导出时间: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
"""

            for product in products:
                # 提取话题 / Extract topics
                topics = product.get('topics', {}).get('edges', [])
                topic_names = [t.get('node', {}).get('name', '') for t in topics if t.get('node')]

                # 提取创作者 / Extract makers
                maker = product.get('maker', {})
                maker_name = maker.get('name', 'Unknown')

                html += f"""
    <div class="product">
        <h2 class="product-name">{product.get('name', '')}</h2>
        <p class="product-tagline">{product.get('tagline', '')}</p>
"""

                if product.get('description'):
                    html += f"""        <p class="product-description">{product.get('description', '')}</p>
"""

                html += f"""        <div class="product-stats">
            <span class="stat">投票: {product.get('votesCount', 0)}</span>
            <span class="stat">评论: {product.get('commentsCount', 0)}</span>
        </div>
        <p class="product-meta">创作者: {maker_name}</p>
"""

                if topic_names:
                    topics_html = ''.join([f'<span class="topic">{name}</span>' for name in topic_names])
                    html += f'        <div class="topics">{topics_html}</div>\n'

                html += f"""        <div class="product-link">
            <a href="{product.get('url', '')}" target="_blank">查看产品</a>
        </div>
    </div>
"""

            html += """</body>
</html>"""

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html)

            print(f"HTML 导出成功: {output_path}")
            return True
        except Exception as e:
            print(f"HTML 导出失败: {e}")
            return False
